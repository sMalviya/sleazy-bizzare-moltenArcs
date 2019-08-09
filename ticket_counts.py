import pandas as pd
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import  Alignment,Font,Border,Side

excel_file = "count_table.xls"

stats_data_frame = pd.read_excel(excel_file)

stats_data_frame['group_final'] = np.where(stats_data_frame['Group'] == 'HTS-GLOBAL-PIB', 'GIS', 'Legacy')
stats_data_frame['row_ct'] = 1
stats_data_frame['Priority'] = np.where(stats_data_frame['Priority'].isin([4,5]), 4, stats_data_frame['Priority'])

df = stats_data_frame[['Priority', 'Status', 'group_final', 'row_ct']]
ser = df.groupby(['group_final', 'Priority', 'Status']).row_ct.sum()

wb = Workbook()

alignment=Alignment(horizontal='general', vertical='center')
font=Font(bold=False)

my_col = openpyxl.styles.colors.Color(rgb='00DCDCDC')
my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_col)
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

sheet1 = wb.active
sheet1.merge_cells('A1:A6')
sheet1.merge_cells('A7:A12')

sheet1.cell(1, 1, 'GSP')
sheet1.cell(7, 1, 'Legacy')

GSP=sheet1.cell(1, 1)
Legacy=sheet1.cell(7,1 )

GSP.font=font
Legacy.font=font
GSP.fill=my_fill
Legacy.fill=my_fill
GSP.alignment = alignment
Legacy.alignment = alignment

sheet1.cell(1, 2, 'Priority')
sheet1.cell(1, 3, 'Acknowledged')
sheet1.cell(1, 4, 'Suspended')
sheet1.cell(1, 5, 'Transferred')
sheet1.cell(1, 6, 'Grand Total')

sheet1.cell(2, 2, 'P1')
sheet1.cell(3, 2, 'P2')
sheet1.cell(4, 2, 'P3')
sheet1.cell(5, 2, 'P4/5')
sheet1.cell(6, 2, 'Total')

sheet1.cell(7, 2, 'Priority')
sheet1.cell(7, 3, 'Acknowledged')
sheet1.cell(7, 4, 'Suspended')
sheet1.cell(7, 5, 'Transferred')
sheet1.cell(7, 6, 'Grand Total')

sheet1.cell(8, 2, 'P1')
sheet1.cell(9, 2, 'P2')
sheet1.cell(10,2, 'P3')
sheet1.cell(11,2, 'P4/5')
sheet1.cell(12, 2, 'Total')

for cell in sheet1["1:1"]:
    cell.font = font
    cell.fill = my_fill

for cell in sheet1["7:7"]:
    cell.font = font
    cell.fill = my_fill

b6 = sheet1['B6']
b6.font = font
b6.fill = my_fill
b12 = sheet1['B12']
b12.font = font
b12.fill = my_fill

BORDER_LIST = ['A1:F1', 'A2:F2', 'A3:F3', 'A4:F4', 'A5:F5', 'A6:F6', 'A7:F7', 'A8:F8', 'A9:F9', 'A10:F10', 'A11:F11', 'A12:F12']

def set_border(ws, cell_range):
    rows = ws[cell_range]
    side = Side(border_style='thin', color="FF000000")

    rows = list(rows)
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side

            # set new border only if it's one of the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border

# border
for pos in BORDER_LIST:
    set_border(sheet1, pos)

for group in ['GIS','Legacy']:
    for p,row in zip(range(1,5),range(2,6)):
        if group == 'Legacy':
            row = row + 6
        ssum=0
        for status, col in zip(['Acknowledged', 'Suspended', 'Transferred'], range(3, 6)):

            try:
                count = ser.loc[group, p, status]
                sheet1.cell(row, col,int(count))
                ssum =int(count)+ssum
                if col==5:
                    sheet1.cell(row, col+1, ssum)
            except:
                    count=0
                    sheet1.cell(row, col, count)
                    if col==5:
                        sheet1.cell(row, col+1, ssum)

for col in range(3, 7):
    csum=0
    for row in range(2,6):
        a=sheet1.cell(row=row,column=col)
        csum+=a.value
        if row==5:
            sheet1.cell(row+1, col, csum)
            csum=0
            row=row+2
            while row <12:
                row=row+1
                a = sheet1.cell(row=row, column=col)
                csum += a.value
                if row==11:
                    sheet1.cell(row + 1, col, csum)
wb.save('out_stats.xlsx')